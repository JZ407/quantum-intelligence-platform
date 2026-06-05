"""
Generate weekly LaTeX/PDF report from MySQL news + conference JSON.

Usage:
    python examples/generate_weekly_report.py --start 2026-05-19 --end 2026-05-25 --issue 42
"""

import os
import sys
import json
import re
import argparse
import subprocess
from datetime import datetime

import pandas as pd
import yaml
from jinja2 import Environment, FileSystemLoader
from sqlalchemy import create_engine

# ------------------------------------------------------------------
# Config
# ------------------------------------------------------------------
DB_URL = 'mysql+pymysql://scraper:scraper123@127.0.0.1:3306/liangke_scraper?charset=utf8mb4'
CATEGORIES = ['宏观态势', '科技前沿', '产品动态', '企业资讯', '资本运作']
TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'weekly_templates')
CONF_JSON_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'conferences_zh.json')

# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def fetch_news(start_date: str, end_date: str):
    """Fetch articles from MySQL for a date range."""
    engine = create_engine(DB_URL)
    query = f"SELECT * FROM articles WHERE liangke_date BETWEEN '{start_date}' AND '{end_date}' ORDER BY id DESC"
    df = pd.read_sql(query, engine)

    def _parse_tags(x):
        if isinstance(x, list):
            return x
        if isinstance(x, str):
            try:
                return json.loads(x)
            except Exception:
                return []
        return []

    df['tags'] = df['tags'].apply(_parse_tags)
    return df


def classify_news(df: pd.DataFrame):
    """Classify articles into 5 categories based on tags.

    Supports both old flat-list format and new dict format with 'weekly' key.
    """
    result = {cat: [] for cat in CATEGORIES}
    for _, row in df.iterrows():
        tags = row.get('tags', [])
        # New format: dict with 'weekly' key containing category labels
        if isinstance(tags, dict):
            weekly_tags = tags.get('weekly', [])
            if isinstance(weekly_tags, list):
                for tag in weekly_tags:
                    if tag in result:
                        result[tag].append(row.to_dict())
                        break
            continue
        # Old format: flat list of category labels
        if not isinstance(tags, list):
            continue
        for tag in tags:
            if tag in result:
                result[tag].append(row.to_dict())
                break
    return result


def preprocess_content(content: str, article_date) -> str:
    """Preprocess article content for weekly report."""
    if not content:
        return ""
    if hasattr(article_date, 'strftime'):
        date_str = article_date.strftime('%Y年%m月%d日')
    else:
        date_str = str(article_date)

    # 1. Clean up leftover punctuation clusters (fuzzy time word removal now handled by LLM)
    content = re.sub(r'[，,]\s*[，,]', '，', content)
    content = re.sub(r'[。．]\s*[，,]', '。', content)
    content = re.sub(r'\s{2,}', ' ', content)

    # 2. Strip existing date prefix with any dash variant, with or without year
    content = re.sub(r'^\s*(\d{4}年)?\s*\d{1,2}月\d{1,2}日\s*[—–‒‐‑―─━－―—\-]+\s*', '', content)
    content = re.sub(r'^\s*(\d{4}年)?\s*\d{1,2}月\d{1,2}日消息[，,]\s*', '', content)
    content = re.sub(r'^\s*\d{4}年\d{1,2}月\d{1,2}日\s*[—–‒‐‑―─━－―—\-]+\s*', '', content)

    # 3. Strip leading year-based date (e.g. "2026年5月11日，")
    content = re.sub(r'^\s*\d{4}年\d{1,2}月(\d{1,2}日|中旬|下旬|上旬)?[，,]\s*', '', content)

    # 4. Strip leading stray punctuation
    content = content.lstrip(' \t\n\r，,。．；;')

    # 5. Add new prefix
    content = f"{date_str}消息，{content}"
    return content


def _load_llm_config():
    cfg_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'config.yaml')
    with open(cfg_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def _get_llm_client():
    sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'rag_system'))
    from llm_client import LLMClient
    cfg = _load_llm_config()['llm']
    return LLMClient(
        provider='openai',
        api_key=cfg['api_key'],
        api_base=cfg['api_base'],
        model=cfg['model'],
        temperature=cfg.get('temperature') if cfg.get('temperature') is not None else 1,
        max_tokens=cfg.get('max_tokens', 2048),
        timeout=180,
    )


def llm_clean_content(categories: dict) -> dict:
    """Use LLM to batch-clean content: remove fuzzy time words, fix date duplication."""
    client = _get_llm_client()

    # Build compact prompt with all articles
    lines = [
        "请检查以下新闻开头，做两件事：",
        "1. 删除'昨日''近日''日前''今天''今早''昨晚''近期''不久前''刚刚'等模糊时间词",
        "2. 如果开头紧挨着'XX月XX日消息，'后又出现类似的日期，删除重复的日期",
        "只修改有明显问题的。输出格式：文章编号|修正后开头",
        "无需修正的也输出原开头。\n"
    ]
    art_index = []
    for cat in CATEGORIES:
        for art in categories[cat]:
            idx = len(art_index) + 1
            art_index.append(art)
            content = art.get('content', '') or ''
            lines.append(f"{idx}|{content[:150]}")

    messages = [
        {"role": "system", "content": "你是新闻编辑助手。只输出要求的格式，不要任何解释。"},
        {"role": "user", "content": "\n".join(lines)},
    ]

    try:
        print('[INFO] LLM cleaning content...')
        resp = client.chat(messages)
        for line in resp.strip().split('\n'):
            line = line.strip()
            if '|' not in line:
                continue
            parts = line.split('|', 1)
            try:
                idx = int(parts[0].strip()) - 1
                if 0 <= idx < len(art_index):
                    fixed = parts[1].strip()
                    if fixed and len(fixed) > 10:
                        orig = art_index[idx].get('content', '') or ''
                        art_index[idx]['content'] = fixed + orig[len(fixed):] if len(orig) > len(fixed) else fixed
            except ValueError:
                continue
    except Exception as e:
        print(f'[WARN] LLM content cleaning failed: {e}')

    # Trim articles to uniform 200-300 chars for weekly digest
    TRIM_THRESHOLD = 300    # trim anything longer than this
    TRIM_TARGET = 250       # target length in Chinese characters
    TRIM_CONTEXT = 2000     # send this many chars to LLM for context
    long_articles = [(i, art) for i, art in enumerate(art_index) if len(art.get('content', '') or '') > TRIM_THRESHOLD]
    if long_articles:
        trim_lines = [
            f"将以下{len(long_articles)}篇新闻精简到200-300字，保留关键数据、人名、机构名、百分比。不要缩成一句话。",
            "输出格式：序号|精简后全文\n"
        ]
        for idx, (gi, art) in enumerate(long_articles, 1):
            content = art.get('content', '') or ''
            trim_lines.append(f"{idx}|{content[:TRIM_CONTEXT]}")
        trim_msg = [
            {"role": "system", "content": "你是新闻编辑，擅长精简。只输出要求的格式。"},
            {"role": "user", "content": "\n".join(trim_lines)},
        ]
        try:
            print(f'[INFO] Trimming {len(long_articles)} long articles via LLM...')
            resp2 = client.chat(trim_msg)
            for line in resp2.strip().split('\n'):
                line = line.strip()
                if '|' not in line:
                    continue
                parts = line.split('|', 1)
                try:
                    tidx = int(parts[0].strip()) - 1
                    if 0 <= tidx < len(long_articles):
                        art_index[long_articles[tidx][0]]['content'] = parts[1].strip()
                except ValueError:
                    continue
        except Exception as e:
            print(f'[WARN] Long article trimming failed: {e}')
    return categories


def generate_summaries_llm(categories: dict) -> dict:
    """Use LLM to generate a paragraph summary per category."""
    client = _get_llm_client()

    prompt_lines = [
        "你是一位量子科技行业分析师。请根据以下本周新闻标题，为每个分类撰写一段概括性摘要。",
        "要求：提炼行业趋势和重点方向，不要写'本周共有X条'，不要简单罗列标题。",
        "每段100-150字。\n"
    ]
    for cat in CATEGORIES:
        arts = categories.get(cat, [])
        if not arts:
            prompt_lines.append(f"\n【{cat}】本周暂无相关新闻。")
            continue
        prompt_lines.append(f"\n【{cat}】（共{len(arts)}条）")
        for art in arts:
            prompt_lines.append(f"- {art['title']}")
    prompt_lines.append("\n请按以下格式输出：")
    prompt_lines.append("宏观态势：<段落>")
    prompt_lines.append("科技前沿：<段落>")
    prompt_lines.append("产品动态：<段落>")
    prompt_lines.append("企业资讯：<段落>")
    prompt_lines.append("资本运作：<段落>")

    messages = [
        {"role": "system", "content": "你是量子科技行业分析师，擅长从大量新闻标题中提炼趋势。"},
        {"role": "user", "content": "\n".join(prompt_lines)},
    ]

    try:
        print('[INFO] Generating paragraph summaries via LLM...')
        resp = client.chat(messages)
        summaries = {}
        for cat in CATEGORIES:
            match = re.search(rf'{re.escape(cat)}[：:](.+?)(?:\n\S|$)', resp, re.DOTALL)
            if match:
                text = match.group(1).strip()
                if len(text) > 20:
                    summaries[cat] = text
                else:
                    summaries[cat] = "本周该领域保持活跃态势。"
            else:
                summaries[cat] = "本周该领域保持活跃态势。"
        return summaries
    except Exception as e:
        print(f'[WARN] LLM summary generation failed: {e}')
        return {cat: "本周该领域保持活跃态势。" for cat in CATEGORIES}


def load_conferences(month: int) -> list:
    """Load conferences for a specific month from JSON cache."""
    if not os.path.exists(CONF_JSON_PATH):
        return []
    with open(CONF_JSON_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    result = []
    for c in data:
        if c.get('month') == month:
            result.append({
                'date_str': c.get('date_str', ''),
                'name': c.get('name_zh', ''),
                'location': c.get('location_zh', ''),
            })
    return result


def render_latex(context: dict, template_name: str = 'weekly_report_template.tex') -> str:
    """Render LaTeX from Jinja2 template."""
    env = Environment(
        loader=FileSystemLoader(TEMPLATE_DIR),
        comment_start_string='{##',
        comment_end_string='##}',
    )
    env.filters['tex_escape'] = tex_escape
    template = env.get_template(template_name)
    return template.render(context)


def tex_escape(s: str) -> str:
    """Escape special LaTeX characters and normalize Unicode."""
    if not isinstance(s, str):
        s = str(s)
    unicode_replacements = [
        ('²', '2'), ('³', '3'), ('⁰', '0'), ('ⁱ', 'i'), ('⁲', '2'), ('⁳', '3'),
        ('⁴', '4'), ('⁵', '5'), ('⁶', '6'), ('⁷', '7'), ('⁸', '8'), ('⁹', '9'),
        ('₀', '0'), ('₁', '1'), ('₂', '2'), ('₃', '3'), ('₄', '4'),
        ('₅', '5'), ('₆', '6'), ('₇', '7'), ('₈', '8'), ('₉', '9'),
    ]
    for old, new in unicode_replacements:
        s = s.replace(old, new)
    replacements = [
        ('\\', '\\textbackslash{}'),
        ('&', '\\&'),
        ('%', '\\%'),
        ('$', '\\$'),
        ('#', '\\#'),
        ('_', '\\_'),
        ('{', '\\{'),
        ('}', '\\}'),
        ('~', '\\textasciitilde{}'),
        ('^', '\\textasciicircum{}'),
    ]
    for old, new in replacements:
        s = s.replace(old, new)
    return s


def compile_pdf(tex_path: str, output_dir: str) -> str:
    """Compile LaTeX to PDF using xelatex (run twice for TOC)."""
    env = os.environ.copy()
    env['PYTHONIOENCODING'] = 'utf-8'
    for i in range(2):
        print(f'  xelatex run {i+1}/2...')
        result = subprocess.run(
            ['xelatex', '-interaction=nonstopmode', '-output-directory', output_dir, os.path.basename(tex_path)],
            cwd=output_dir,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            env=env,
        )
        out = result.stdout.decode('utf-8', errors='replace') if result.stdout else ''
        err = result.stderr.decode('utf-8', errors='replace') if result.stderr else ''
        if result.returncode != 0:
            print(f'[WARN] xelatex returned {result.returncode}')
            if '!' in out:
                print(out[-2000:].encode('gbk', errors='replace').decode('gbk', errors='replace'))
    pdf_path = tex_path.replace('.tex', '.pdf')
    return pdf_path


def _find_col(df: pd.DataFrame, candidates: list):
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in cols_lower:
            return cols_lower[cand.lower()]
    return None


def _parse_tender_df(df: pd.DataFrame) -> list:
    cn = _find_col(df, ['项目名称', '项目标题', 'title', 'name'])
    cd = _find_col(df, ['项目基本情况', '项目内容', 'desc', 'description'])
    cu = _find_col(df, ['项目主体单位', '采购单位', 'unit', 'organization'])
    cs = _find_col(df, ['项目规模', '预算金额', 'scale', 'budget'])
    cp = _find_col(df, ['发布时间', '发布日期', 'pub_date'])
    cr = _find_col(df, ['预采时间', '预计采购时间', 'pre_date'])
    curl = _find_col(df, ['信息来源', '来源链接', 'url', 'source', 'link'])
    results = []
    for _, row in df.iterrows():
        t = {
            'name': str(row[cn]) if cn and pd.notna(row[cn]) else '',
            'desc': str(row[cd]) if cd and pd.notna(row[cd]) else '',
            'unit': str(row[cu]) if cu and pd.notna(row[cu]) else '',
            'scale': str(row[cs]) if cs and pd.notna(row[cs]) else '',
            'pub_date': str(row[cp]) if cp and pd.notna(row[cp]) else '',
            'pre_date': str(row[cr]) if cr and pd.notna(row[cr]) else '',
            'url': str(row[curl]) if curl and pd.notna(row[curl]) else '',
        }
        t = {k: (v if v not in ('nan', 'None', 'null') else '') for k, v in t.items()}
        results.append(t)
    return results


def _filter_patents_llm(patents: list, category: str, max_keep: int = 8) -> list:
    """LLM ranks patents, prioritizing 授权 > 公开 > 实用新型 > 外观."""
    if len(patents) <= max_keep:
        return patents
    client = _get_llm_client()
    lines = [
        f"以下是【{category}】板块的{len(patents)}条专利。请筛选出最重要的{max_keep}条放入周报。",
        "优先级：授权专利 > 公开专利 > 实用新型专利 > 外观专利。在同类中，优先创新型、突破性技术。",
        "只输出选中的序号（逗号分隔），不要解释。\n"
    ]
    for i, p in enumerate(patents, 1):
        lines.append(f"{i}. [{p.get('type','未知')}] {p.get('title','')} | {p.get('applicant','')} | {p.get('abstract','')[:80]}")
    messages = [
        {"role": "system", "content": "你是专利分析专家。"},
        {"role": "user", "content": "\n".join(lines)},
    ]
    try:
        resp = client.chat(messages)
        nums = re.findall(r'\d+', resp)
        selected = []
        for n in nums:
            idx = int(n) - 1
            if 0 <= idx < len(patents) and patents[idx] not in selected:
                selected.append(patents[idx])
        return selected[:max_keep] if selected else patents[:max_keep]
    except Exception:
        return patents[:max_keep]


def _filter_patents_by_date(patents: list, start_date: str, end_date: str) -> list:
    """Filter patents to only those with dates within the given range."""
    from datetime import datetime as _dt
    sd = _dt.strptime(start_date, '%Y-%m-%d').date()
    ed = _dt.strptime(end_date, '%Y-%m-%d').date()
    result = []
    for p in patents:
        d = p.get('date', '')
        if not d:
            result.append(p)
            continue
        try:
            for fmt in ['%Y-%m-%d', '%Y%m%d', '%Y/%m/%d', '%Y年%m月%d日']:
                try:
                    if sd <= _dt.strptime(str(d).strip()[:10], fmt).date() <= ed:
                        result.append(p)
                    break
                except ValueError:
                    continue
        except Exception:
            result.append(p)
    return result


def _parse_patent_df(df: pd.DataFrame, excel_path: str = None) -> list:
    ct = _find_col(df, ['标题(译)(简体中文)', '发明名称(中文)(机器翻译)', '标题', 'title'])
    ca = _find_col(df, ['[标]当前申请(专利权)人', '申请人', 'applicant'])
    ci = _find_col(df, ['发明人', 'inventor'])
    cd = _find_col(df, ['公开(公告)日', '公开日', '授权日', 'date'])
    cab = _find_col(df, ['摘要(译)(简体中文)', '摘要(中文)(机器翻译)', '摘要', 'abstract'])
    curl = _find_col(df, ['url', '专利链接', '链接'])
    ctype = _find_col(df, ['法律状态/事件', '法律状态', '专利类型', 'type'])

    # Extract hyperlinks from 公开号 column
    hyperlinks = {}
    if excel_path:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            num_col_idx = None
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=1, column=col).value
                if v and '公开' in str(v) and '号' in str(v):
                    num_col_idx = col
                    break
            if num_col_idx:
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=num_col_idx)
                    if cell.hyperlink and cell.hyperlink.target:
                        hyperlinks[row_idx - 2] = cell.hyperlink.target
            wb.close()
        except Exception:
            pass

    results = []
    for i, (_, row) in enumerate(df.iterrows()):
        ptype = str(row[ctype]) if ctype and pd.notna(row[ctype]) else ''
        if '授权' in ptype:
            ptype = '授权专利'
        elif '公开' in ptype:
            ptype = '公开专利'
        else:
            ptype = '公开专利'
        url = hyperlinks.get(i, '') or (str(row[curl]) if curl and pd.notna(row[curl]) else '')
        if url in ('nan', 'None', 'null'):
            url = ''
        p = {
            'title': str(row[ct]) if ct and pd.notna(row[ct]) else '',
            'applicant': str(row[ca]) if ca and pd.notna(row[ca]) else '',
            'inventor': str(row[ci]) if ci and pd.notna(row[ci]) else '',
            'type': ptype,
            'date': str(row[cd]) if cd and pd.notna(row[cd]) else '',
            'abstract': str(row[cab]) if cab and pd.notna(row[cab]) else '',
            'url': url,
        }
        p = {k: (v if v not in ('nan', 'None', 'null') else '') for k, v in p.items()}
        # Remove Japanese kana and Korean hangul that break xelatex CJK fonts
        for key in ('title', 'applicant', 'inventor', 'abstract'):
            p[key] = re.sub(r'[぀-ヿ가-힯ᄀ-ᇿ]+', ' ', p[key])
            p[key] = re.sub(r'\s+', ' ', p[key]).strip()
        results.append(p)
    return results


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Generate weekly quantum news report')
    parser.add_argument('--start', required=True, help='Start date (YYYY-MM-DD)')
    parser.add_argument('--end', required=True, help='End date (YYYY-MM-DD)')
    parser.add_argument('--issue', required=True, help='Issue number (e.g. 42)')
    parser.add_argument('--output', default=None, help='Output directory')
    parser.add_argument('--conf-month', type=int, default=None, help='Conference month to include')
    parser.add_argument('--tender-excel', default=None, help='Path to tender Excel file')
    parser.add_argument('--patent-excel', action='append', default=[], help='Path to patent Excel file (repeatable)')
    parser.add_argument('--page-types', default='flash,article,reference',
                        help='Comma-separated page types to include (default: flash,article,reference)')
    parser.add_argument('--selected-ids', default='',
                        help='Comma-separated article IDs to include (overrides other filters)')
    args = parser.parse_args()

    output_dir = args.output or os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'weekly_output')
    os.makedirs(output_dir, exist_ok=True)

    # 1. Fetch and classify news
    print(f'[INFO] Fetching news from {args.start} to {args.end}...')
    df = fetch_news(args.start, args.end)
    print(f'[INFO] Fetched {len(df)} articles')

    # 1.5. Apply page_type filter
    page_types = [t.strip() for t in args.page_types.split(',') if t.strip()]
    if page_types and 'page_type' in df.columns:
        df = df[df['page_type'].isin(page_types)]
        print(f'[INFO] After page_type filter ({args.page_types}): {len(df)} articles')

    # 1.6. Apply selected-ids filter (from UI checkboxes)
    if args.selected_ids:
        selected = set(int(x) for x in args.selected_ids.split(',') if x.strip())
        df = df[df['id'].astype(int).isin(selected)]
        print(f'[INFO] After selected-ids filter: {len(df)} articles')

    categories = classify_news(df)
    for cat, arts in categories.items():
        print(f'  {cat}: {len(arts)} articles')

    # 2. Preprocess article content (regex: date prefix removal) and set URL
    for cat in CATEGORIES:
        for art in categories[cat]:
            art['content'] = preprocess_content(art.get('content', '') or '', art.get('liangke_date', ''))
            art['url'] = art.get('reference_url') or art.get('liangke_url') or ''

    # 3. LLM content review: remove fuzzy time words, fix date duplication
    categories = llm_clean_content(categories)

    # 4. Generate paragraph summaries via LLM
    summaries = generate_summaries_llm(categories)

    # 5. Conferences (from JSON cache)
    conferences = []
    if args.conf_month:
        print(f'[INFO] Loading conferences for month {args.conf_month} from cache...')
        conferences = load_conferences(args.conf_month)
        print(f'[INFO] {len(conferences)} conferences found')

    # 5. Optional: tenders from Excel
    tenders = []
    if args.tender_excel and os.path.exists(args.tender_excel):
        print(f'[INFO] Reading tenders from {args.tender_excel}...')
        tdf = pd.read_excel(args.tender_excel)
        tenders = _parse_tender_df(tdf)
        print(f'[INFO] Loaded {len(tenders)} tenders')

    # 6. Optional: patents from Excel
    patents = []
    for pexcel in args.patent_excel:
        if os.path.exists(pexcel):
            print(f'[INFO] Reading patents from {pexcel}...')
            pdf = pd.read_excel(pexcel)
            group = _parse_patent_df(pdf, excel_path=pexcel)
            # Infer category from filename
            PATENT_CATS = ['低温环境系统', '超导量子测控技术', '量子软件与算法', '量子算力网', '量子科技长三角产业创新中心']
            cat = '其他'
            for c in PATENT_CATS:
                if c in pexcel:
                    cat = c
                    break
            if group:
                group = _filter_patents_by_date(group, args.start, args.end)
                if len(group) > 8:
                    group = _filter_patents_llm(group, cat)
                print(f'[INFO] {cat}: {len(group)} patents (after date + LLM filter)')
                if group:
                    patents.append({'category': cat, 'entries': group})

    # 7. Escape article content for LaTeX before rendering
    for cat in CATEGORIES:
        for art in categories[cat]:
            for key in ['title', 'content', 'url']:
                if key in art and art[key]:
                    art[key] = tex_escape(art[key])
        summaries[cat] = tex_escape(summaries[cat])

    for c in conferences:
        c['name'] = tex_escape(c.get('name', ''))
        c['location'] = tex_escape(c.get('location', ''))

    # 8. Build patent summary for abstract
    patent_summary = ''
    if patents:
        total_p = sum(len(g.get('entries', g.get('items', []))) for g in patents)
        cats_p = '、'.join(g['category'] for g in patents)
        patent_summary = f"本周专利动态涵盖{cats_p}共{len(patents)}个板块、{total_p}条专利。"

    # 9. Render LaTeX
    print('[INFO] Rendering LaTeX...')
    context = {
        'start_date': args.start,
        'end_date': args.end,
        'issue_no': args.issue,
        'summaries': summaries,
        'categories': categories,
        'conferences': conferences,
        'conf_month': args.conf_month,
        'tenders': tenders,
        'patents': patents,
        'patent_summary': patent_summary,
    }
    latex = render_latex(context)

    tex_path = os.path.join(output_dir, f'量子行业每周新闻洞察_第{args.issue}期.tex')
    with open(tex_path, 'w', encoding='utf-8') as f:
        f.write(latex)
    print(f'[OK] LaTeX saved to {tex_path}')

    # 9. Copy cover image
    cover_candidates = [
        os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'weekly_templates', 'Cover_Suzhou.png'),
    ]
    for cover_src in cover_candidates:
        if os.path.exists(cover_src):
            import shutil
            shutil.copy(cover_src, os.path.join(output_dir, 'Cover_Suzhou.png'))
            print(f'[OK] Cover image copied from {cover_src}')
            break
    else:
        print(f'[WARN] Cover image not found')

    # 10. Compile PDF
    print('[INFO] Compiling PDF (xelatex x2)...')
    pdf_path = compile_pdf(tex_path, output_dir)
    print(f'[OK] PDF saved to {pdf_path}')


if __name__ == '__main__':
    main()
